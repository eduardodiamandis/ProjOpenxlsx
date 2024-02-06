from openpyxl import Workbook
wb = Workbook()


ws=wb.active

ws1 = wb.create_chartsheet("Minha planilha") #cria uma tabela  vazia com o nome "Minha planilha"

ws2 = wb.create_chartsheet("Automação")


def mostra_planilhas():
    """Retorna x, nomes de planilhas em wb"""
    for sheets in wb:
        x = wb.sheetnames
        print(x)

def backup_planilha(nome_planilha,dest):
    #Para atualizar
    
    nome_planilha = wb.active
    wb.save(filename="Backup/"+dest+"-"+str(datetime.now())+".xlsx")
    
#adiciona dados na linha  e colunas
c = ws['A4']
c = 4
i = 20
while i >= 10:
    i-=1
    print(ws[i])


